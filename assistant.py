import time
import os
import hashlib
import csv
from colorama import Fore,init
from utils.helperfunctio import load_study_history,save_study_session,load_notes,save_note
from utils.helperfunctio import load_assignment,save_assignment
from utils.helperfunctio import total_study_sessions,total_study_time,total_notes_count,total_assignments_count,pending_assignments,completed_assignments
from utils.helperfunctio import load_users,save_users,save_goals,load_goals,save_pomodoro,load_pomodoro,load_streaks,save_streak,load_tasks,save_tasks,load_activity,save_activity
from utils.helperfunctio import high_priority_tasks,low_priority_tasks,medium_priority_tasks,total_tasks,completed_tasks,pending_tasks,highprior_andnotcompleted,nothighandnotcompleted
from datetime import datetime,timedelta
import matplotlib.pyplot as plt
from reportlab.platypus import *
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    PageBreak
)
from openpyxl import Workbook
import zipfile
from utils.helperfunctio import load_Settings,save_settings
init(autoreset=True)
current_user=None
def pasrse_date(date_string):
    return datetime.strptime(date_string,"%d/%m/%Y")





def study_timer(current_username):
    load=load_Settings(current_username)
    while(True):
     try:
      minutes=int(input(Fore.CYAN+"Enter how many minutes you want to study or press 0 to use  default study time : "))
      break
     except ValueError:
        print(Fore.RED+"Invalid input, press 0 to use default study time")
    if minutes==0:
        minutes=load[0]['default_study_timer']
    print(f"{Fore.YELLOW}study session started for {minutes} minutes")
    time.sleep(minutes)
    print(Fore.GREEN+"congratulations !")
    print(Fore.GREEN+f"you have completed your study session {minutes} minutes")
    history=load_study_history(current_username)
    today=datetime.now().date()
    strtime=today.strftime("%d/%m/%Y")
    streak=False
    for his in history:
        # print("running")
        if not his['date']==strtime:
            streak=True
            # print("running")
             
    if streak==True:
      update_streaks(current_username)
    #   print("running")

    history.append({
        "duration":minutes,
        "date":strtime
    })
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"📚 Started study session"
    })
    save_activity(load_act,current_username)
    
    
    save_study_session(history,current_username)
    
def view_study_history(current_username):
    history=load_study_history(current_username)
    if not history:
        print("\n no study session found")
        return
    print(f"\n{Fore.BLUE}===== STUDY HISTORY =====")
    for i, session in enumerate(history,start=1):
        print(f"{i}. {session["duration"]} minutes")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"⭐ Viewed study history"
    })
    save_activity(load_act,current_username)




def add_assignment(current_username):
    assignments=load_assignment(current_username)
    title=input(Fore.CYAN+"enter asssignment title : ")
    while(True):
     due_date=input(Fore.CYAN+"enter due date of assignment (DD/MM/YYYY) : ")
     
     try:
        datetime.strptime(due_date,"%d/%m/%Y")
        break
     except ValueError:
        print(f"{due_date} invalid format ")

    
    assignments.append({
        "title":title,
        "due_date":due_date,
        "completed":False
    })
    save_assignment(assignments,current_username)
    print(f"{Fore.GREEN} assignment added successfully")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"➕ Added {title} assignment"
    })
    save_activity(load_act,current_username)
def view_assignment(current_username):
    assignments=load_assignment(current_username)
    if not assignments:
        print("no assignment found")
        return
    print(f"\n{Fore.BLUE}===== ASSIGNMENTS =====")
    for i,assignment in enumerate(assignments,start=1):
        status="completed✅" if assignment["completed"] else "pending❌"
        print(f"{i}. {assignment['title']} | DUE:{assignment['due_date']} | STATUS:{status}")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"⭐ Viewed assignments"
    })
    save_activity(load_act,current_username)
# def view_not_completed_assignments(current_username):
#     assignments=load_assignment(current_username)
#     if not assignments:
#         print("no assignment found")
#         return
#     print(f"\n{Fore.BLUE}===== NOT COMPLETED ASSIGNMENTS =====")
#     for i,assignment in enumerate(assignments,start=1):
#         if not assignment["completed"]:
#             status="pending❌"
#             print(f"{i}. {assignment['title']} | DUE:{assignment['due_date']} | STATUS:{status}")
def mark_complete(current_username):
    assignments=load_assignment(current_username)
    if not assignments:
        print("no assignment found")
        return
    view_assignment(current_username)
    while(True):
     try: 
        num=int(input(Fore.CYAN+"Enter number which assignment you want to mark as completed : "))
        if num>=1 and num<=len(assignments):
            assignments[num-1]["completed"]=True
            save_assignment(assignments,current_username)
            break
        else:
            print("invalid number")
        
     except ValueError:
        print("invalid number") 
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"☑️ Completed assignment {assignments[num-1]['title']}"
    })
    save_activity(load_act,current_username)
    # while(True):    
    #     verify=input(Fore.RED+"Are you sure you want to mark this assignment as completed? (yes/no): ")
    
    #     if verify.lower()=='yes':
    #         print("assignment marked as completed")
    #         break
    #     elif verify.lower()=='no':
    #         print("marking cancelled")
    #         return
    #     else:
    #         print("invalid choice")
    
     
def delete_Assignment(current_username):
    assignments=load_assignment(current_username)
    if not assignments:
        print("no assignments found")
        return
    view_assignment(current_username)

    while(True):
     try:
        num=int(input(Fore.CYAN+"Enter number which assignment you want to delete : "))
        break
     except ValueError:
        print("invalid number")
    while(True):
      verify=input(Fore.RED+"Are you sure you want to delete this assignment? (yes/no): ")
      if verify.lower()=='yes':
            if num>=1 and num<=len(assignments) and verify.lower()=='yes':
                assignments.pop(num-1)
                save_assignment(assignments,current_username)
                print("assignment deleted")
                break
      elif verify.lower()=='no':
            print("deletion cancelled")
            return
      else:
            print("invalid choice")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"✔️ Deleted {assignments[num-1]['title']} assignment "
    })
    save_activity(load_act,current_username)

            
        
    
def assignment_menu():
    global current_user
    while(True):
        print(f"\n{Fore.BLUE}===== ASSIGNMENT MENU =====")
        print(f"{Fore.MAGENTA}1. Add Assignment")
        print(f"{Fore.MAGENTA}2. View Assignment")
        print(f"{Fore.MAGENTA}3. Mark assignment as completed")
        print(f"{Fore.MAGENTA}4. Delete Assignment")
        print(f"{Fore.RED}5. Back")

        choice = input(Fore.CYAN+"Enter choice: ")

        if choice == "1":
            add_assignment(current_user)

        elif choice == "2":
            view_assignment(current_user)

        elif choice == "3":
            mark_complete(current_user)

        elif choice == "4":
            delete_Assignment(current_user)

        elif choice == "5":
            break

        else:
            print("Invalid choice.")
def add_note(current_username):
    note=load_notes(current_username)
    new_note=input(Fore.CYAN+"enter new note you want to add")
    note.append(new_note)
    save_note(note,current_username)
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"📔 Note Added"
    })
    save_activity(load_act,current_username)
def view_notes(current_username):
    note=load_notes(current_username)
    if not note:
        print("no note found")
        return
    print(f"\n{Fore.BLUE}===== NOTES =====")
    for i,notes in enumerate(note,start=1):
        print(f"{i}. {notes}")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"⭐ Viewed notes"
    })
    save_activity(load_act,current_username)
    
def delete_note(current_username):
    note=load_notes(current_username)
    if not note:
       print("no notes to be deleted")
       return 
    try:
      view_notes(current_username)
      num=int(input(Fore.CYAN+"\nEnter number which note you want to delete"))
      
      if num>=1 and num<=len(note):
        deleted=note.pop(num-1)
        save_note(note,current_username)
        print(f"deleted {deleted}")
      else:
        print("invalid note number")
    except ValueError:
        print("invalid number ")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":" 📃 Note Deleted"
    })
    save_activity(load_act,current_username)
def notes_menu():
    global current_user
    while(True):
        print(f"\n{Fore.BLUE}===== NOTES MENU =====")
        print(f"{Fore.MAGENTA}1. Add Note")
        print(f"{Fore.MAGENTA}2. View Notes")
        print(f"{Fore.MAGENTA}3. Delete Note")
        print(f"{Fore.RED}4. Back")

        choice = input(Fore.CYAN+"Enter choice: ")

        if choice == "1":
            add_note(current_user)

        elif choice == "2":
            view_notes(current_user)

        elif choice == "3":
            delete_note(current_user)

        elif choice == "4":
            break

        else:
            print(Fore.RED+"Invalid choice.")
def export_notes(current_username):
            notes=load_notes(current_username)
            with open(f"exports/{current_username}/notes.csv",'w',newline='') as file:
                writer=csv.writer(file)
                writer.writerow(['note'])
                for note in notes:
                    writer.writerow([note])
            print(f"{Fore.GREEN} Notes exported successfully to exports/notes.csv")
            load_act=load_activity(current_username)
            timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
            load_act.append({
        "time":timedat,
        "activity":"📝 Exported notes"
    })
            save_activity(load_act,current_username)
def assignment_export(current_username):
    assignments=load_assignment(current_username)
    with open(f"exports/{current_username}/assignment.csv",'w') as file:
        writer=csv.writer(file)
        writer.writerow(['Title','Due_Date','Completed'])
        for assignment in assignments:
            writer.writerow([assignment['title'],assignment['due_date'],assignment['completed']])
    print(f"{Fore.GREEN} Assignment exported successfully to exports/{current_username}assignment.csv")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"☑️ Assignment Exported"
    })
    save_activity(load_act,current_username)
def export_study_history(current_username):
    history=load_study_history(current_username)
    with open(f"exports/{current_username}/study_history.csv",'w',newline='')as file:
        writer=csv.writer(file)
        writer.writerow(['date','duration'])
        for session in history:
            writer.writerow([session['date'],session['duration']])
    print(f"{Fore.GREEN} Study history exported successfully to exports/study_history.csv")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"✔️ Exported Study history"
    })
    save_activity(load_act,current_username)
def export_tasks(current_username):
    load=load_tasks(current_username)
    file=f"exports/{current_username}/tasks.csv"
    with open(file,'w',newline='')as fi:
       writer=csv.writer(fi)
       writer.writerow(['Title','Priority','Deadline','Completed'])
       for task in load:
           writer.writerow([task['Title'],task['Priority'],task['Deadline'],task['Completed']])
    print(f"{Fore.GREEN} Tasks exported successfully to exports/{current_username}/study_history.csv")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"☑️ Tasks Exported"
    })
    save_activity(load_act,current_username)

        
        
def export_menu():
    global current_user
    while(True):
        print(f"\n{Fore.BLUE}===== EXPORT MENU =====")
        print(f"{Fore.MAGENTA}1. export notes")
        print(f"{Fore.MAGENTA}2. export assignment")
        print(f"{Fore.MAGENTA}3. export study history")
        print(f"{Fore.MAGENTA}4. export Tasks")
        print(f"{Fore.RED}5. Back")
        choice = input(Fore.CYAN+"Enter choice: ")
        if choice == "1":
            export_notes(current_user)
        elif choice == '2':
            assignment_export(current_user)
        elif choice == '3':
            export_study_history(current_user)
        elif choice=='4':
            export_tasks(current_user)
        elif choice == "5":
            break
        else:
            print("Invalid choice.")
def overdue_count(current_username):
    load=load_assignment(current_username)
    today=datetime.now().date()
    overdues_count=0
    for assign in load:
        if pasrse_date(assign['due_date']).date() < today and not assign['completed']:
            overdues_count+=1
    return overdues_count

def set_goals(current_username):
    load=[]
    goal=int(input(f"{Fore.CYAN}Enter daily study goal: "))
    today=datetime.now().date()
    load.append({"goals":goal})
    save_goals(load,current_username)
    print(Fore.GREEN+"Goal Saved")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":" ➕ Added Goal"
    })
    save_activity(load_act,current_username)
def view_goal(current_username):
    load=load_goals(current_username)
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"⭐ Viewed Goals"
    })
    save_activity(load_act,current_username)
    if not load:
        return Fore.RED+"please add goal..."
    return int(load[0]['goals'])



def pomodoro(current_username):
    setting=load_Settings(current_username)
    study_minutes=int(input(Fore.CYAN+"Enter how many minutes you want to study or press enter for default  : ") or setting[0]['pomodoro'])
    break_time=int(input(Fore.CYAN+"Enter how many minutes of break you want  or default 5 : ")or load[0]['break_time'])
    total_seconds=study_minutes
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"Started Pomodoro"
    })
    save_activity(load_act,current_username)
    while(total_seconds):
        minutes=total_seconds//60
        sec=total_seconds% 60
        print(f"{Fore.YELLOW}{minutes:02d}:{sec:02d}")
        time.sleep(1)
        total_seconds-=1
    print(Fore.LIGHTBLUE_EX+"\n\nWell done study session completed!")
    history=load_study_history(current_username)
    today=datetime.now().date()
    strtime=today.strftime("%d/%m/%Y")
    streak=False
    for his in history:
        # print("running")
        if not his['date']==strtime:
            # print("running")
            streak=True
    if streak==True:
    #   print("running")
      update_streaks(current_username)
    
    history.append({
        "duration":study_minutes,
        "date":strtime
    })
    save_study_session(history,current_username)
    load=load_pomodoro(current_username)
    load[0]['pomodoro']+=1
    print(load)
   
    
    save_pomodoro(load,current_username)
    print(f"\n{Fore.GREEN} Take a break for {break_time} minutes")
    break_seco=break_time
    while break_seco:
        mins=break_seco//60
        secs=break_seco%60
        print(f"{Fore.YELLOW}{mins:02d}:{secs:02d}")
        time.sleep(1)
        break_seco-=1

    print(f"{Fore.RED} Break over!")

    continu=input(Fore.CYAN+"Do you want to continue or exit pomodoro timer Yes or No: ")
    if continu.lower()=='yes':
        pomodoro(current_username)
    else:return
# pomodoro(current_user)

def view_progress(current_username):
    load=load_study_history(current_username)
    today=datetime.now().date()
    strintime=today.strftime("%d/%m/%Y")
    study_min=0
    goal=load_goals(current_username)
    goa=goal[0]['goals']
    print(type(goa))
    for duration in load:
        # print(duration['date'])
        if duration['date']==strintime:
            study_min+=duration['duration']
            # print(type(study_min))
    percentage=(study_min / goa)*100
    print(f"{Fore.YELLOW}Goal:{goa}mins")
    print(f"{Fore.YELLOW}completed :{study_min}mins")
    print(f"{Fore.YELLOW}Progress: {percentage:.1f}%")
    load_act=load_activity(current_username)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"⭐ Viewed Progress"
    })
    save_activity(load_act,current_username)
    
    if percentage>=100:
        print(f"{Fore.GREEN}Goal achieved!")
def progress(current_username):
    load=load_study_history(current_username)
    today=datetime.now().date()
    strintime=today.strftime("%d/%m/%Y")
    study_min=0
    goal=load_goals(current_username)
    if not goal:
         return Fore.RED+"please add goal in productivity features"
    goa=goal[0]['goals']
    # print(type(goa))
    for duration in load:
        # print(duration['date'])
        if duration['date']==strintime:
            study_min+=duration['duration']
            # print(type(study_min))
    percentage=(study_min / goa)*100
    # print(f"{Fore.YELLOW}Goal:{goa}mins")
    print(f"{Fore.YELLOW}completed :{Fore.WHITE}{study_min}mins")
    print(f"{Fore.YELLOW}Progress: {Fore.WHITE}{percentage:.1f}%")
    return
    


def productivity_menu():
    global current_user
    
    while(True):
        print(Fore.BLUE+"======  Productivity Menu  ======")
        print(f"{Fore.MAGENTA}1. Set goal")
        print(f"{Fore.MAGENTA}2. Start pomodoro")
        print(f"{Fore.MAGENTA}3. View Progress")
        print(f"{Fore.MAGENTA}4. Exit") 
        choice=input(Fore.CYAN+"enter your choice : ")
        if choice=='1':
            set_goals(current_user)
        elif choice=='2':

            pomodoro(current_user)
        elif choice=='3':
            view_progress(current_user)
        elif choice=="4":
            break
        else:
            print(Fore.RED+"Invalid choice")
    
def progress_bar(percentage):
    filled = int(percentage // 10)
    empty = 10 - filled

    return "█" * filled + "░" * empty


def search_notes(current_username):
    note=load_notes(current_username)
    if not note:
        print("no notes found")
        return
    keyword=input(Fore.CYAN+"enter keyword to search in notes")
    result=[n for n in note if keyword.lower() in n.lower()]
    if result:
        print(f"{Fore.GREEN} FOUND {len(result)} notes containing this {keyword}")
        for i,notes in enumerate(result,start=1):
            print(f"{i}. {notes}")
    else:
        print(Fore.RED+"no keyword match to notes")
def search_assignments(current_username):
    data=load_assignment(current_username)
    keyword=input(Fore.CYAN+"enter title of assignment to search in assignments")
    result=[n for n in data if keyword.lower() in n['title'].lower()]
    if result:
        print(f"{Fore.GREEN} found {len(keyword)} matches in assignment")
        for i,assi in enumerate(result,start=1):
            print(f"""{i}. Title : {assi['title']}
                           DUE: {assi['due_date']}
                           Status: {assi['status']}   """)
    else:
        print(f"{Fore.RED}No keyword match the assignment title")
def pending1_assignments(current_username):
    assign=load_assignment(current_username)
    for a in assign:
        if not a['completed']:
            print(f"{Fore.RED}Pending Assignment: {a['title']} | DUE: {a['due_date']}")
def completed1_assignmments(current_username):
    assign=load_assignment(current_username)
    for a in assign:
        if a['completed']:
            print(f"{Fore.GREEN}Completed Assignment: {a['title']} | DUE: {a['due_date']}")
def search_tasks(current_username):
    load=load_tasks(current_username)
    keyword=input("Enter task title to search in tasks: ")
    result=[n for n in load if keyword.lower()in n['Title'].lower()]
    if result:
        print(f"{Fore.MAGENTA}==== Matching keywords ====")
        for k in result:
            print(f"""Title: {k['Title']}
Priority: {k['Priority']}
Deadline: {k['Deadline']}
Completed: {k['Completed']}""")
   
def search_menu():
  global current_user

  while(True):
    print(f"\n{Fore.BLUE}===== SEARCH MENU =====")
    print(f"{Fore.MAGENTA} 1. Seacrh notes")
    print(f"{Fore.MAGENTA} 2. Search assignments")
    print(f"{Fore.MAGENTA} 3. Search Tasks")
    print(f"{Fore.MAGENTA} 4. VIEW PENDING ASSIGNMENTS")
    print(f"{Fore.MAGENTA} 5. VIEW COMPLETED ASSIGNMENTS")
    print(f"{Fore.RED} 6. Back")
    choice=input(Fore.CYAN+"Enter your choice: ")
    if choice=="1":
        search_notes(current_user)
    elif choice=='2':
        search_assignments(current_user)
    elif choice=='3':
        search_tasks(current_user)
    elif choice=='4':
        pending1_assignments(current_user)
    elif choice=='5':
        completed1_assignmments(current_user)
    elif choice=="6":
        print(f"{Fore.BLUE} returning to main menu...")
        break
    else:
        print(f"{Fore.RED}invalid choice")

def overdue_assignment(current_username):
    print(Fore.BLUE+"====== Overdue assignments ======")
    load=load_assignment(current_username)
    today=datetime.now().date()
    found=False
    i=1
    for assign in load:
        if pasrse_date(assign['due_date']).date()<today and not assign['completed']:
            print("\n")
            print(f"{Fore.RED}{i}. title :{assign['title']}")
            print(f"{Fore.RED}     assignment is due on {assign['due_date']}")
            found=True
            i+=1

    if found==False:
     print(Fore.GREEN+"No assignments overdue")
    

def due_assignments(current_username):
    load=load_assignment(current_username)
    today=datetime.now().date()
    found=False
    print(f"{Fore.BLUE} ====== DUE ASSIGNMENTS ====== ")
    i=1
    for assign in load:       
        if pasrse_date(assign['due_date']).date()> today and not assign['completed']:
            print("\n")
            print(f"{Fore.YELLOW}{i}. Title : {assign['title']}")
            print(f"{Fore.YELLOW}     Due date: {assign['due_date']}")
            found=True
            i+=1
    if not found:
        print(Fore.GREEN+"No assignments due")
def upcoming_assignments(currren_username):
    print(Fore.YELLOW+" ====== Upcomming assignments in 3 days ======")
    i=1
    load=load_assignment(currren_username)
    today=datetime.now().date()
    next_days=today + timedelta(days=3)
    found=False
    for assign in load:
        if today<pasrse_date(assign['due_date']).date()<next_days:
            print("\n")
            print(Fore.LIGHTBLUE_EX+f"{i}. Title : {assign['title']}")
            print(Fore.LIGHTBLUE_EX+f"     Due_date {assign['due_date']}")
            found=True
            i=+1

    if not found:
        print(Fore.GREEN+"No upcoming assignment in 3 Days")
        print("\n")

def alerts_menu():
   global current_user
   while(True):
    print(Fore.MAGENTA+"1. Check overdue assignments")
    print(Fore.MAGENTA+"2. Check Due assignments")
    print(Fore.MAGENTA+"3. Check Upcoming assignments")
    print(Fore.MAGENTA+"4. Exit")
    Choice=input(Fore.CYAN+"Enter your choice : ")
    if Choice=='1':
        overdue_assignment(current_user)
    elif Choice=='2':
        due_assignments(current_user)
    elif Choice=='3':
        upcoming_assignments(current_user)
    elif Choice=='4':
        print(Fore.GREEN+"Exiting alerts")
        break
    else:
        print(Fore.RED+"invalid choice ")


def hashpass(password):
    return hashlib.sha256(password.encode()).hexdigest()
def create_user_files(username):
    os.mkdir(
        f"data/{username}"
    )

    files=['assignment.json','notes.json','tasks.json','study_history.json','study_goals.json','pomodoro.json','streak.json','badges.json','activity.json']
    expor=['assignment.csv','notes.csv','studyhistory.csv','tasks.csv']
    for file in files:
        with open(f'data/{username}/{file}','w')as fi:
            if file=='pomodoro.json':
                fi.write('[{"pomodoro":0}]')
            elif file=='streak.json':
                fi.write('[{"streaks":0,"yesterday":""}]')
            else:
              fi.write("[]")
    os.mkdir(f"exports/{username}")


def register():
  load=load_users()
  username=input(f"{Fore.MAGENTA} Enter username: ")
  password=hashpass(input(f"{Fore.MAGENTA} Enter password: "))
  
  for user in load:
      if user['username']==username:
        print(f"{Fore.RED} User Already Exists!")
        return
  load.append({
      'username':username,
      'password':password})    
  save_users(load)
  print(Fore.GREEN+"Registration Successfully!")
  create_user_files(username)




  


def login():
    load=load_users()
    username=input(f"{Fore.MAGENTA} Enter username: ")
    password=hashpass(input(f"{Fore.MAGENTA} Enter password: "))
    for user in load:
        if user['username']==username and user['password']==password:
            global current_user
            current_user=username
            print(Fore.GREEN+"login successfully") 
            setting=load_Settings(current_user)
            print(f"\n{Fore.YELLOW}Welcome : {setting[0]['display_name']}")
            load_act=load_activity(current_user)
            timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
            load_act.append({
        "time":timedat,
        "activity":"🔗 logined in account"
    })
            save_activity(load_act,current_user)
            return True
    print(Fore.RED+"Invalid crendtials")    
    return False   
def profile():
    global current_user
    
    print(Fore.BLUE+"\n ======== PROFILE  ========")
    print(f"username :  {current_user}")
def logout():
    global current_user
    current_user=None
    print(f"{Fore.GREEN} Successfully logout")
def change_password():
    global current_user
    load=load_users()
    for user in load:
        if user['username']==current_user:
            new_password=input(Fore.CYAN+"Enter new password : ")
            user['password']=new_password
            save_users(load)
            print(f"{Fore.GREEN} Password changed successfully")
            load_act=load_activity(current_user)
            timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
            load_act.append({
        "time":timedat,
        "activity":"🔐 Changed password"
    })
            save_activity(load_act,current_user)

def change_username():
    load=load_users()
    global current_user
    new_username=input(Fore.CYAN+"Enter new  ")
    for name in load:
            if name['username']==new_username:
                print("username already exists")
                return
    for user in load:
       if  user['username']==current_user:
           os.rename(f"exports/{current_user}",f"exports/{new_username}")
           os.rename(f"data/{current_user}",f"data/{new_username}")
           os.rename(f"backups/{current_user}_backup.zip",f"backups/{new_username}_backup.zip")

           user['username']=new_username
           save_users(load)

           current_user=new_username

           print(f"{Fore.GREEN} Username changed successfully")
           load_act=load_activity(current_user)
           timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
           load_act.append({
        "time":timedat,
        "activity":"☑️ Changed Username"
    })
           save_activity(load_act,current_user)

def delete_account():
    global current_user
    load=load_users()
    confirm=input("Type DELETE to confirm deletion of account")
    if confirm.lower()!='delete':
        print(Fore.GREEN+"deletion cancelled")
        return
    load=[user for user in load if user['username']!=current_user]
    save_users(load)
    print(Fore.YELLOW+"Account Deleted Succesfully")
    current_user=None
    return True
def account_details():

    global current_user

    print(Fore.BLUE+"\n===== ACCOUNT =====")
    
    print(
        f"{Fore.GREEN}Username: {current_user}"
    )
def setting_menu():
  global current_user
  while(True):
    load=load_Settings(current_user)
    print(Fore.BLUE+"====== Settings ======")
    print(Fore.MAGENTA+"1. Change username ")
    print(Fore.MAGENTA+"2. Change Password ")
    print(Fore.MAGENTA+"3. Delete account")
    print(Fore.MAGENTA+"4. View account details")
    print(Fore.MAGENTA+"5. Change dispaly name")
    print(Fore.MAGENTA+"6. change default study")
    print(Fore.MAGENTA+"7. change default pomodoro timer")
    print(Fore.MAGENTA+"8. change default break time in pomodoro")
    print(Fore.MAGENTA+"9. change notifications settings")

    print(Fore.MAGENTA+"10. exit")
    choice=input(Fore.CYAN+"Enter your choice")
    if choice=='1':
        change_username()
    elif choice=='2':
        change_password()
    elif choice=='3':
        if delete_account():
            return "logout"
    elif choice=='4':
        account_details()
    elif choice=='5':
        load[0]['display_name']=input("Enter display name : ")
    elif choice=='6':
        load[0]['default_study_timer']=input("Enter study time you want to set as default : ")
    elif choice=='7':
        load[0]['pomodoro']=input("Enter pomodoro timer you want to set as default : ")
    elif choice=='8':
        load[0]['break_time']=input("Enter break time you want to set as default : ")
    elif choice=='9':
       while(True):
        notification=input("Do you want notifications in dashboard yes or no ")
        if notification.lower()=='yes':
           load[0]['show_notifications']=True
           print(Fore.GREEN+"Notification settings changed sucessfully")

        elif notification.lower=='no':
           load[0]['show_notifications']=False
           print(Fore.GREEN+"Notification settings changed sucessfully")

        else:
            print("Invalid choice")
            
        
    elif choice=='10':
        break
    else:
        print(Fore.RED+"Invaid choice")


def add_tasks(current_username):
    load=load_tasks(current_username)
    title=input("Enter title of the task: ")
    while(True):
     priority=input(Fore.CYAN+"Enter priority of the task high,low ,medium: " + Fore.WHITE)
     if priority.lower() in ['high','medium','low']:
        break
     else:
         print(Fore.RED+"Invalid priority")
    completed=False
    while(True):
     try:
      deadline=input(Fore.CYAN+'Enter deadline of the task (%d/%m/%y): ' + Fore.WHITE)
      datetime.strptime(deadline,"%d/%m/%Y")
      break
     except ValueError:
      print(Fore.RED+"Invalid date format")
      
    load.append({
        "Title":title,
        "Priority":priority,
        "Completed":completed,
        "Deadline":deadline
    })
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"➕  {title} Task added"
    })
    save_activity(load_act,current_user)
    save_tasks(load,current_username)
def view_tasks(current_username):
    load=load_tasks(current_username)
    for i,tasks in enumerate(load,start=1):
        status="completed✅" if tasks['Completed'] else "Pending"
        print(Fore.YELLOW+f"{i}. Title:{tasks['Title']} \n Priority:{tasks['Priority']} \n {tasks['Deadline']} \n Status:{status}")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"⭐ Viewed tasks"
    })
    save_activity(load_act,current_user)
        
def delete_tasks(current_username):
    view_tasks(current_username)
    load=load_tasks(current_username)

    while(True):
      try:
        choice=int(input(Fore.CYAN+"Enter which task you want to delete " ))
        
      except ValueError:
        print("invalid choice")
     
      if choice>=1 and choice<=len(load):
           verify=input(Fore.RED+"Are you sure you want to delete this task? (yes/no): ")
           if verify.lower()=='yes':
             load.pop(choice-1)
             save_tasks(load,current_username)
             print(Fore.GREEN+"tasks deleted sucessfully")
             load_act=load_activity(current_user)
             timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
             load_act.append({
        "time":timedat,
        "activity":f"{load[choice-1]['title']} task deleted"
    })
             save_activity(load_act,current_user)
             break
    
           elif verify.lower()=='no':
             print(Fore.GREEN+"deletion cancelled")
             break
           else:
             print(Fore.RED+"invalid choice")
      else:
        print(Fore.RED+"invalid choice")
        
           
     
def mark(current_username):
    load=load_tasks(current_username)
    if not load:
        print("no tasks found")
        return
    view_tasks(current_username)
    while(True):
     try: 
        choice=int(input(Fore.CYAN+"Enter no which task you to mark as completed : "))
     except ValueError:
        print("invalid choice")
     if choice>=1 and choice<=len(load):
          load[choice-1]['Completed']=True
          save_tasks(load,current_username)
          load_act=load_activity(current_user)
          timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
          load_act.append({
        "time":timedat,
        "activity":f"✔️  Mark {load[choice-1]['title']} ccompleted"
    })
          save_activity(load_act,current_user)
          break
     else:
          print("invalid choice")
     
def filter_task(current_username):
    load=load_tasks(current_username)
    if not load:
        return
    high=[]
    low=[]
    medium=[]
    pending=[]
    completed=[]
    for task in load:
        if task['Priority'].lower()=='high':
            high.append(task)
        elif task['Priority'].lower()=='medium':
            medium.append(task)
        elif task['Priority']=='low':
            low.append(task)
        if task['Completed']:
            completed.append(task)
        else:
            pending.append(task)
    while(True):
      print("======  Filter menu  ======")
      print(Fore.MAGENTA+"1. Filter tasks basis on high priority")
      print(Fore.MAGENTA+"2. Filter tasks basis on medium priority")
      print(Fore.MAGENTA+"3. Filter tasks basis on low priority")
      print(Fore.MAGENTA+"4. Filter tasks basis on completed priority")
      print(Fore.MAGENTA+"5. Filter tasks basis on pending priority")
      print(Fore.MAGENTA+"6. exit the filter feature")
      choice=input(Fore.CYAN+"Please enter how you want filter to task : ")
      if choice=='1' or choice.lower()=='high':
         print("===== High priority tasks =====")
         if not high:
              print(Fore.RED+"No high priority tasks")
         else:
          for i,tasks in enumerate(high,start=1):
              status="completed✅" if tasks['Completed'] else "Pending"
              print(Fore.YELLOW+f"{i}. Title:{tasks['Title']} \n Priority:{tasks['Priority']} \n {tasks['Deadline']} \n Status:{status}")
      elif choice=='2' or choice.lower()=='medium':
         print("===== Medium priority tasks =====")
         if not medium:
              print(Fore.RED+"No Medium priority tasks")
         else:
          for i,tasks in enumerate(medium,start=1):
              status="completed✅" if tasks['Completed'] else "Pending"
              print(Fore.YELLOW+f"{i}. Title:{tasks['Title']} \n Priority:{tasks['Priority']} \n {tasks['Deadline']} \n Status:{status}")
      elif choice=='3' or choice.lower()=='low':
         print("===== Low priority tasks =====")
         if not low:
          print(Fore.RED+"NO low priority tasks")
         else:

          for i,tasks in enumerate(low,start=1):
              status="completed✅" if tasks['Completed'] else "Pending"
              print(Fore.YELLOW+f"{i}. Title:{tasks['Title']} \n Priority:{tasks['Priority']} \n {tasks['Deadline']} \n Status:{status}")
      elif choice=='4' or choice.lower()=='pending':
         print("===== Completed tasks =====")
         if not completed:
              print(Fore.RED+"No Completed tasks")
         else:
          for i,tasks in enumerate(completed,start=1):
              status="completed✅" if tasks['Completed'] else "Pending"
              print(Fore.YELLOW+f"{i}. Title:{tasks['Title']} \n Priority:{tasks['Priority']} \n {tasks['Deadline']} \n Status:{status}")
      elif choice=='5' or choice.lower()=='pending':
        print("===== Pending tasks =====")
        if not pending:
              print(Fore.RED+"No Pending tasks")
        else:
          for i,tasks in enumerate(pending,start=1):
              status="completed✅" if tasks['Completed'] else "Pending"
              print(Fore.YELLOW+f"{i}. Title:{tasks['Title']} \n Priority:{tasks['Priority']} \n {tasks['Deadline']} \n Status:{status}")
      elif choice=='6'or choice=='back':
          break
      else:
          print("invalid value")
          
def tasks(current_username):
    while(True):
       print(Fore.BLUE+"=======  Tasks menu  =======")
       print(Fore.MAGENTA+"1. Add tasks")
       print(Fore.MAGENTA+"2. Delete tasks")
       print(Fore.MAGENTA+"3. View tasks")
       print(Fore.MAGENTA+"4. Mark tasks complete")
       print(Fore.MAGENTA+"5. Filter tasks")
       print(Fore.MAGENTA+"6. Back")
       choice=input("Enter your choice: ")
       if choice=='1':
           add_tasks(current_username)
       elif choice=='2':
           delete_tasks(current_username)
       elif choice=='3':
           view_tasks(current_username)
       elif choice=='4':
           mark(current_username)
       elif choice=='5':
           filter_task(current_username)
       elif choice=='6':
           break
       else:
           print(Fore.RED+"Invalid choice")






    


def productivity_score(current_username):
    load=load_assignment(current_username)
    streak=load_streaks(current_username)
    completed=sum(1 for a in load if a['completed'])
    score=completed*10+streak[0]['streaks']*5
    return  score

def week_recommendation(current_username):
    load_assignmen=load_assignment(current_username)
    load_study=load_study_history(current_username)
    today=datetime.now().date()
    daybef=(today-timedelta(days=4)).strftime("%d/%m/%Y")
    minutes=0
    for stud in load_study:
        if stud['date']>daybef:
            minutes+=stud['duration']
    pomo=load_pomodoro(current_username)
    goals=load_goals(current_username)
    pending=sum(1 for a  in load_assignmen if not a['completed'])
    print(f"======== Recommendations 💫 ========")
    if minutes<300:
        print(Fore.YELLOW+"📊 Increase study time this week")
    if pending>5:
        print(Fore.RED+"⚠️  You have many pening assignments")
    if not goals:
        print(Fore.YELLOW+"You have no study goals please add goals ")
    if  not pomo:
        print(Fore.YELLOW+"You have no pomodoro sessins, please do to increase your study time")
    elif pomo[0]['pomodoro']<5:
        print(Fore.YELLOW+"You have less pomodoro sessins, please do to increase your study time")
def notification(current_username):
    load=load_assignment(current_username)
    pending=sum(1 for a in load if a['completed'])
    print(Fore.BLUE+"======== Notifications ========")
    print(f"{Fore.RED} You have {highprior_andnotcompleted(current_username)} High Priority pending tasks ")
    # print(f"other pending tasks are : {nothighandnotcompleted(current_username)}")
    today=datetime.now().date().strftime("%d/%m/%Y")
    Today_assign=0
    load=load_assignment(current_username)
    for assign in load:
        if assign['due_date']==today and not assign['completed']:
            Today_assign+=1
    if Today_assign>0:
        print(f"{Fore.RED} You have {Today_assign} assignments due today")
    overdue=overdue_count(current_username)
    if overdue>0:
        print(f"{Fore.RED} You have {overdue} overdue assignments")
    loadd=load_study_history(current_username)
    for his in loadd:
        if his['date']!=today:
            print(f"{Fore.RED} You have not studied today")
            break


    print(f"{Fore.RED} You have {pending} pending assignmemnts")
    Streak=load_streaks(current_username)
    print(f"{Fore.YELLOW} Current streak {Streak[0]['streaks']}")

def study_alerts(current_username):
     load=load_study_history(current_username)
     today=datetime.now().date()
     if not load :
        #  print(f"{Fore.YELLOW} You have not studied today, please start studying")
         return True
     for his in load:
         if his['date']!=today.strftime("%d/%m/%y"):
             return True
     return False
def smart_alerts(current_username):
    load=load_assignment(current_username)
    for assign in load:
        if assign['due_date']>(datetime.now().date()).strftime("%d/%m/%Y") and not assign['completed']:
            print(f"{Fore.RED} {assign['title']} assignment is over due")
        elif assign['due_date']==(datetime.now().date()).strftime("%d/%m/%Y") and not assign['completed']:
            print(f"{Fore.YELLOW} {assign['title']} is due today")
    high=highprior_andnotcompleted(current_username)
    if high>0:
        print(f"{Fore.RED} You have {high} high priority pending tasks")    
    if study_alerts(current_username):
        print(f"{Fore.RED} You have not studied today, please start studying")

def recommendation(current_username):
    load=load_assignment(current_username)
    today=datetime.now().date()
    first_assign=[]
    second_assign=[]
    one_day_due_assign=[]
    first_task=[]
    second_task=[]
    one_day_due_task=[]
    print(f"{Fore.BLUE} ====== Recommendations 💫 ======")
    if load:
     for assign in load:
        if assign['due_date']<today.strftime("%d/%m/%Y") and not assign['completed']:
           first_assign.append(assign)
        elif assign['due_date']==today.strftime("%d/%m/%Y") and not assign['completed']:
          second_assign.append(assign)
        elif assign['due_date']==(today+timedelta(days=1)).strftime("%d/%m/%Y") and not assign['completed']:
            one_day_due_assign.append(assign)
        
    else:
        print(Fore.RED+"No Assignments added found")
        
    tasks=load_tasks(current_username)

    for task in tasks:

        if task['Priority'].lower()=='high' and not task['Completed'] and task['Deadline']<today.strftime("%d/%m/%Y"):
            first_task.append(task)
        if task['Priority'].lower()=='medium' and not task['Completed'] and task['Deadline']<today.strftime("%d/%m/%Y"):  
            first_task.append(task)
        # elif task['Priority'].lower()=='low' and not task['Completed'] and task['Deadline']<today.strftime("%d/%m/%Y"):
            # first_task.append(task)
        if task['Priority'].lower()=='high' and not task['Completed'] and task['Deadline']==today.strftime("%d/%m/%Y"):
            second_task.append(task)
        if task['Priority'].lower()=='medium' and not task['Completed'] and task['Deadline']==today.strftime("%d/%m/%Y"):  
            second_task.append(task)
        # elif task['Priority'].lower()=='low' and not task['Completed'] and task['Deadline']==today.strftime("%d/%m/%Y"):
            # second_task.append(task)
        if task['Priority'].lower()=='high' and not task['Completed'] and task['Deadline']==(today+timedelta(days=1)).strftime("%d/%m/%Y"):
            one_day_due_task.append(task)
        if task['Priority'].lower()=='medium' and not task['Completed'] and task['Deadline']==(today+timedelta(days=1)).strftime("%d/%m/%Y"):  
            one_day_due_task.append(task)
        # elif task['Priority'].lower()=='low' and not task['Completed'] and task['Deadline']==(today+timedelta(days=1)).strftime("%d/%m/%Y"):
            # one_day_due_task.append(task)
    st=0
    if first_assign:
     for i,assign in enumerate(first_assign,start=1):
        print(f"{Fore.RED}{i}. Do {assign['title']} assignment is overdue")
        st+=1
    if first_task:
     for i,task in enumerate(first_task,start=st):
        print(f"{Fore.YELLOW}{i}. Do {task['Title']} because it is {task['Priority']} Priority and overdue")
        st+=1
    if second_assign:
     for i,assign in enumerate(second_assign,start=st):
        print(f"{Fore.YELLOW}{i}. Do {assign['title']} assignment is due today")
        st+=1
    if second_task:
     for i,task in enumerate(second_task,start=st):
        print(f"{Fore.YELLOW}{i}. Do {task['Title']} because it is {task['Priority']} Priority and due today")
        st+=1
    if one_day_due_assign:
     for i,assign in enumerate(one_day_due_assign,start=st):
        print(f"{Fore.GREEN}{i}. Do {assign['title']} assignment is due tomorrow")
        st+=1
    if one_day_due_task:
     for i,task in enumerate(one_day_due_task,start=st):
        print(f"{Fore.GREEN}{i}. Do {task['Title']} because it is {task['Priority']} Priority and due tomorrow")
        st+=1
   
    load_assignmen=load_assignment(current_username)
    load_study=load_study_history(current_username)
    today=datetime.now().date()
    daybef=(today-timedelta(days=4)).strftime("%d/%m/%Y")
    minutes=0
    for stud in load_study:
        if stud['date']>daybef:
            minutes+=stud['duration']
    pomo=load_pomodoro(current_username)
    goals=load_goals(current_username)
    pending=sum(1 for a  in load_assignmen if not a['completed'])
    
    if minutes<300:
        print(Fore.YELLOW+"📊 Increase study time this week")
    if pending>5:
        print(Fore.RED+"⚠️  You have many pening assignments")
    if not goals:
        print(Fore.YELLOW+"You have no study goals please add goals ")
    if  not pomo:
        print(Fore.YELLOW+"You have no pomodoro sessins, please do to increase your study time")
    elif pomo[0]['pomodoro']<5:
        print(Fore.YELLOW+"You have less pomodoro sessins, please do to increase your study time")
# recommendation(current_user)
     
            
def weekly_summary(current_username):
    load_assignmen=load_assignment(current_username)
    load_study=load_study_history(current_username)
    today=datetime.now().date()
    daybef=(today-timedelta(days=7)).strftime("%d/%m/%Y")
    min=0
    completed=0
    pending=0
    session=0
    for his in load_study:
        if his['date']>daybef:
            min+=his['duration']
            session+=1

    for assign in load_assignmen:
        if assign['completed']:
                completed+=1
        else:
                pending+=1
    print(Fore.BLUE+"====== Weekly summary ======")
    print(Fore.YELLOW+f"Total week study sessions :  {session}")
    print(Fore.YELLOW+f"Total week study minutes :  {min}")
    print(f"{Fore.GREEN} Completed assignments : {completed}")
    print(f"{Fore.RED} Pending assignments : {pending}") 
def insights_menu():
    global current_user
    while(True):
        print(Fore.MAGENTA+"1. Productivity score")    
        print(Fore.MAGENTA+"2. Recommendations")        
        print(Fore.MAGENTA+"3. Notifications")        
        print(Fore.MAGENTA+"4. Weekly summary")        
        print(Fore.MAGENTA+"5. Back")        
        choice=input("Enter your choice : ")
        if choice=='1':
            print(f"{Fore.YELLOW}productivity_score: {productivity_score(current_user)}")
        elif choice=='2':
            recommendation(current_user)
        elif choice=='3':
            notification(current_user)
        elif choice=='4':
            weekly_summary(current_user)
        elif choice=='5':
            break
        else:
            print(Fore.RED+"Invalid choice")

def pdf_Report(current_username):
    file_path=f"data/{current_username}/productivity_Report.pdf"
    pdf=SimpleDocTemplate(file_path)
    styles=getSampleStyleSheet()
    content=[]
   
    title1=Paragraph(
        """<b>
        =====
        Ai Student Assistant report
        =====</b>""",
        styles["Title"]
    )
    content.append(title1)
    content.append(Spacer(1,12))
    content.append(Paragraph(f"<b>Generated on : </b>{(datetime.now().date()).strftime("%d/%m/%Y")}",styles['BodyText']))
    content.append(Spacer(1,10))
    content.append(Paragraph(f"<b>username: </b>{current_username} ",styles['BodyText']))
    content.append(Spacer(1,10))
    title2=Paragraph(
        """<b>
        ===========
              STUDY STATISTICS
        ===========</b>""",
        styles["Title"]
    )
    content.append(title2)
    content.append(Spacer(1,12))
    history=load_study_history(current_username)
    assignments=load_assignment(current_username)
    Streak=load_streaks(current_username)
    study_time=sum(session['duration'] for session in history)
    completed=sum(1 for a in assignments if a['completed'])
    pending=len(assignments)-completed
    Streaks=Streak[0]['streaks']

    pdf_text=[
       

       f"<b>Total Study Time: </b>{study_time}mins",

       f"<b> Total Study Sessions: </b>{len(history)}"
  
       f"<b>Current Streak: </b>{Streaks}",
       
    ]
    for con in pdf_text:
      content.append(Paragraph(con,styles['BodyText']))
      content.append(Spacer(1,10))
    choice=input(Fore.MAGENTA+"Do you want graph and charts in pdf Yes or No: ")
    if choice.lower()=='yes':
        study_graph(current_username)
        assignment_chart(current_username)
        content.append(Image(f"data/{current_username}/study_graph.png",
                         width=400,
                         height=250))
        content.append(PageBreak())
    title3=Paragraph(
        f"""<b>




       ========
        ASSIGNMENT SUMMARY
       ========</b>""",
        styles["Title"]
    )
    content.append(title3)
    content.append(Spacer(1,12))
    Assign_text=[
       
      f"<b>Completed Assignments: </b>{completed}",
       
       f"<b>Pending Assignments:</b>{pending}",
       
    ]
    content.append(Paragraph(Assign_text[0],styles['BodyText']))
    content.append(Spacer(1,10))
    content.append(Paragraph(Assign_text[1],styles['BodyText']))
    content.append(Spacer(1,10))
    if choice.lower()=='yes':
       content.append(Image(f"data/{current_username}/assignment_chart.png",
                         width=350,
                         height=250))
    title6=Paragraph( f"""<b>




       ========
        TASKS SUMMARY
       ========</b>""",styles['Title'])
    content.append(title6)
    content.append(Spacer(1,10))
    task_text=[f"Total tasks: {total_tasks(current_username)}",f"Completed Task : {completed_tasks(current_username)}",f"Pending Task: {pending_tasks(current_username)}",f"High Priority tasks: {high_priority_tasks(current_username)}"]
    for text in task_text:
        content.append(Paragraph(text,styles['BodyText']))
        content.append(Spacer(1,10))
    title4=Paragraph(
        """<b>
       -------------
                  PRODUCTIVITY SCORE
       ------------</b>""",
        styles["Title"]
    )
    content.append(title4)
    content.append(Spacer(1,12))
    score=productivity_score(current_username)
    content.append(Paragraph(f"Productivity Score: {score}",styles['BodyText']))
    content.append(Spacer(1,10))
    title5=Paragraph(
        """<b>
       -------------
               ACHIEVEMENTS 
       -------------
       </b>""",
        styles["Title"]
    )
    content.append(title5)
    content.append(Spacer(1,12))
    loa=get_badges(current_username)
    
    for badge in loa:
        content.append(Paragraph(f"{badge}"))
        content.append(Spacer(1,10))
    title1=Paragraph(
        """<b>
       ==================================
        Generated by Ai Student Assistant
       ==================================</b>""",
        styles["Title"]
    )
    content.append(title1)
    content.append(Spacer(1,12))
    
    pdf.build(content)
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"Pdf report generated"
    })
    save_activity(load_act,current_user)
    print(Fore.GREEN+"📑 Pdf Report Generated")
    print(f"Saved at : {file_path}")
    
    
def import_tasks(current_username):
    load=[]
    with open(f"exports/{current_username}/tasks.csv",'r') as file:
        read=csv.DictReader(file)
        for row in read:
            load.append(row)
    save_tasks(load,current_username)
    print(Fore.GREEN+"Tasks imported sucessfully")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"📃 Tasks Imported"
    })
    save_activity(load_act,current_user)

def import_notes(current_username):
    load=[]
    with open(f"export/{current_username}/notes.csv") as file:
        read=csv.DictReader(file)
        for note in read:
            load.append(note)
    save_note(load,current_username)
    print(Fore.GREEN+"Notes imported sucessfully")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"🗒️ Notes imported"
    })
    save_activity(load_act,current_user)

def import_assignments(current_username):
    load=[]
    with open(f"exports/{current_username}/assignment.csv") as f:
        read=csv.DictReader(f)
        for assign in read:
            load.append(assign)
    save_assignment(load,current_username)
    print(Fore.GREEN+"assignment imported sucessfully")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"📝 Assignments imported"
    })
    save_activity(load_act,current_user)

def import_study(current_username):
    load=[]
    with open(f"exports/{current_username}/study_history.csv") as f:
        read=csv.DictReader(f)
        for study in read:
            load.append(study)
    save_study_session(load,current_username)
    print(Fore.GREEN+"Study imported sucessfully")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"📔 Study history imported"
    })
    save_activity(load_act,current_user)
def import_menu():
    global current_user
    while(True):
        print(Fore.MAGENTA+"1. Import tasks")
        print(Fore.MAGENTA+"2. Import notes")
        print(Fore.MAGENTA+"3. Import assignment")
        print(Fore.MAGENTA+"4. Import study sessions")
        print(Fore.MAGENTA+"5. Back")
        choice=input(Fore.CYAN+"Enter your choice")
        if choice=='1':
            import_tasks(current_user)
        elif choice=='2':
            import_notes(current_user)
        elif choice=='3':
            import_assignments(current_user)
        elif choice=='4':
            import_study(current_user)
        elif choice=='5':
            break
        else:
            print(Fore.RED+"Invalid choice")

        
    

        


        

            

def auth():
  while(True):
    print(f"{Fore.BLUE}======= Login system =======")
    print(f"{Fore.MAGENTA}1. Register")
    print(f"{Fore.MAGENTA}2. Login")
    print(f"{Fore.MAGENTA}3. Exit")
    choice=input(f"{Fore.CYAN} Enter your choice : ")
    if choice=='1':
        register()
    elif choice=='2':
        if login():
            global current_user
            load=load_Settings(current_user)
            if load[0]['show_notifications']:
                notification(current_user)
            return True
    elif choice=='3':
        return False
    else:
        print(Fore.RED+"Invalid choice")
def study_graph(current_username):
    
    load_study=load_study_history(current_username)
    size=len(load_study)
    sessions=list(range(1,size+1))
    minutes=[]
    for his in load_study:
        minutes.append(his['duration'])
    plt.figure(figsize=(8,5))
    plt.plot(sessions,minutes,marker='o')
    plt.title("study session")
    plt.ylabel("minutes")
    plt.xlabel("sessions")
    plt.grid(True)
    plt.savefig(f"data/{current_username}/study_graph.png")
    print(f"Study graph  uploaded in data/{current_username}/study_graph.png")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"💹 Study graph uploaded"
    })
    save_activity(load_act,current_user)
# study_graph()
def assignment_chart(current_username):
    
    assignments=load_assignment(current_username)
    completed=sum(1 for x in assignments if x['completed']==True)
    # print(completed)
    pending=len(assignments)-completed
    # print(pending)
    labels=['completed','pending']
    values=[completed,pending]
    plt.figure(figsize=(6,6))
    plt.pie(values,labels=labels,autopct="%1.1f%%")
    plt.title("assignment chart")
    # plt.show()
    plt.savefig(f"data/{current_username}/assignment_chart.png")
    print(f"Assignment chart uploaded in data/{current_username}/assignment_chart.png")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"📝 Uploaded assignment chart"
    })
    save_activity(load_act,current_user)



def excel(current_username):
    wb=Workbook()
    sheet=wb.active
    sheet.title="Study history"
    sheet.append([
        "Date",
        "Duration"
    ]) 
    total_time=0
    history=load_study_history(current_username)
    for his in history:
        total_time+=his['duration']
        sheet.append([
            his['date'],
            his['duration']
        ])
    assignment_sheet=wb.create_sheet("Assignments")
    assignment_sheet.append([
            "Title",
            "Due_date",
            "Completed"
        ])
    assign=load_assignment(current_username)
    for ass in assign:
            assignment_sheet.append(
[ass['title'],
 ass['due_date'],
 ass['completed']]
            )
    tasks=wb.create_sheet("Tasks")
    tasks.append(["Title","Priority","completed","Deadline"])
    load_task=load_tasks(current_username)
    for tas in load_task:
        tasks.append([tas['Title'],tas['Priority'],tas['Completed'],tas['Deadline']])
    analytics=wb.create_sheet("Analytics")
    analytics.append([
            "Metric","Value"
        ])
    analytics.append(["study_sessons",len(history)])
    analytics.append([
    "Study Time",
    total_time
])
    streak=load_streaks(current_username)

    analytics.append([
    "Current Streak",
    streak[0]['streaks']
])
    score=productivity_score(current_username)
    analytics.append([
    "Productivity Score",
     score
])
    file_path = (
    f"data/{current_user}/"
    "productivity_report.xlsx"
)

    wb.save(file_path)
    print(Fore.GREEN+"Excel report Generated ✅")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":"📃 Excel report generated"
    })
    save_activity(load_act,current_user)
def reports_menu():
   global current_user
   while(True):
    print(Fore.BLUE+"====== Reports ======")
    print(Fore.MAGENTA+"1. Generate Study graph")
    print(Fore.MAGENTA+"2. Generate Assignment chart")
    print(Fore.MAGENTA+"3. Generate Pdf Report")
    print(Fore.MAGENTA+"4. Generate Excel Report")
    print(Fore.MAGENTA+"5. Back")
    choice=input(Fore.CYAN+"Enter your choice: ")
    if choice=='1':
        study_graph(current_user)
    elif choice=='2':
       assignment_chart(current_user)
    elif choice=='3':
        pdf_Report(current_user)
    elif choice=='4':
        excel(current_user)
    elif choice=='5':
        break
    else:
        print(Fore.RED+"Invalid choice")
    

def create_backup(current_username):
    source_folder=f"data/{current_username}"
    backupfolder=f"backups/{current_username}_backup.zip"
    with zipfile.ZipFile(backupfolder,'w',zipfile.ZIP_DEFLATED) as zipf:
        for root,dirs,files in os.walk(source_folder):
          for file in files:
            path=os.path.join(root,file)
            arcname=os.path.relpath(path,source_folder)
            zipf.write(path,arcname)
    print(Fore.GREEN+"backup created succesfully")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"📦 Backup created"
    })
    save_activity(load_act,current_user)
def restore_backup(current_username):
    target_file=f"data/{current_username}"
    backup_file=f"backups/{current_username}_backup.zip"
    if not os.path.exists(backup_file):
        print(f"{Fore.RED}backup_file not existed")
        return
    with zipfile.ZipFile(backup_file,'r') as zipf:
        zipf.extractall(target_file)
    print(Fore.GREEN+"backup restored")
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    load_act.append({
        "time":timedat,
        "activity":f"📦  Backup restored"
    })
    save_activity(load_act,current_user)
def backup_menu():
    global current_user
    print(Fore.BLUE+"======== BACKUP SYSTEM ========")
    while(True):
        print(Fore.MAGENTA+"1. Create backup")
        print(Fore.MAGENTA+"2. Restore backup")
        print(Fore.MAGENTA+"3. back")
        choice=input(Fore.CYAN+"Enter your choice : ")
        if choice=="1":
            create_backup(current_user)
        elif choice=="2":
            restore_backup(current_user)
        elif choice=="3":
            break
        else:
            print(Fore.RED+"Invalid choice")
import random
def motivational_quotes():
    quotes = [
    "Discipline beats motivation.",
    "Small progress is still progress.",
    "Success starts with consistency.",
    "Study today, succeed tomorrow.",
    "Every expert was once a beginner.",
    "Stay focused and never give up.",
    "Dream big, work hard.",
    "One chapter at a time.",
    "Learning never goes to waste.",
    "Keep showing up every day.",
    "Hard work always pays off.",
    "Your future is created today.",
    "Progress, not perfection.",
    "Success is earned, not given.",
    "Believe in your potential.",
    "Keep moving forward.",
    "Make today count.",
    "Consistency creates excellence.",
    "Small efforts lead to big results.",
    "Don't stop until you're proud.",
    "Knowledge is your greatest investment.",
    "Learn something new every day.",
    "Stay hungry for knowledge.",
    "Focus on improvement.",
    "Be stronger than your excuses.",
    "Push yourself beyond limits.",
    "You are capable of amazing things.",
    "The best investment is in yourself.",
    "Action beats intention.",
    "Success loves preparation.",
    "Every day is a new opportunity.",
    "Your habits shape your future.",
    "Study with purpose.",
    "A little progress every day adds up.",
    "Be patient with your growth.",
    "Work in silence, let success speak.",
    "Stay disciplined even when it's hard.",
    "You don't have to be perfect, just consistent.",
    "Keep learning, keep growing.",
    "Failure is a lesson, not the end.",
    "Great things take time.",
    "Never stop believing in yourself.",
    "Every minute of study counts.",
    "Start now, not later.",
    "One task at a time.",
    "Focus on what matters.",
    "Turn dreams into goals.",
    "Goals require action.",
    "Learn, practice, improve.",
    "Success is built daily.",
    "Your future self will thank you.",
    "The pain of discipline is temporary.",
    "Keep your eyes on the goal.",
    "Stay positive and productive.",
    "Confidence comes from preparation.",
    "Discipline creates freedom.",
    "Never underestimate small efforts.",
    "Keep building your future.",
    "Success begins with self-belief.",
    "Today's effort is tomorrow's success.",
    "Stay committed to your goals.",
    "The journey is worth it.",
    "Learn from yesterday, improve today.",
    "Consistency is your superpower.",
    "You are stronger than your doubts.",
    "Every accomplishment starts with a decision.",
    "Focus creates results.",
    "Work hard, stay humble.",
    "Keep your momentum alive.",
    "Nothing changes unless you do.",
    "You can do difficult things.",
    "Success comes to those who prepare.",
    "Do something today your future self will appreciate.",
    "Build habits that build success.",
    "Great achievements begin with small steps.",
    "Never quit on your dreams.",
    "Stay curious, keep learning.",
    "Every challenge makes you stronger.",
    "Success is a series of small wins.",
    "Your effort matters.",
    "Every page you read brings you closer to success.",
    "Learn with passion.",
    "Believe, achieve, repeat.",
    "Keep improving every single day.",
    "Your only competition is yesterday's you.",
    "Knowledge opens every door.",
    "One more hour of effort can change everything.",
    "Winners never stop learning.",
    "The harder you work, the luckier you become.",
    "Stay dedicated to your purpose.",
    "Make discipline your lifestyle.",
    "Success begins outside your comfort zone.",
    "The secret to success is consistency.",
    "Invest in your mind every day.",
    "Every study session brings you closer to your goals.",
    "Your dreams deserve your effort.",
    "The best project you'll ever work on is yourself.",
    "Stay focused, stay determined.",
    "Growth happens one day at a time.",
    "Keep chasing excellence.",
    "Believe in the process.",
    "Your journey has just begun."
]
    return random.choice(quotes)
def update_streaks(current_username):
     load=load_streaks(current_username)
     print(load)
     today=datetime.now().date()
     strintime=today.strftime('%d/%m/%Y')
     yesterday=today-timedelta(days=1)
     yesterdaystrin=yesterday.strftime("%d/%m/%Y")
     if load[0]['yesterday']==yesterdaystrin or load[0]['yesterday']=="":
         load[0]['streaks']+=1
         load[0]['yesterday']=strintime
         save_streak(current_username,load)
def view_streak(current_username):
    load=load_streaks(current_username)
    return load[0]['streaks']
def get_badges(current_username):
    badges=[]
    history=load_study_history(current_username)
    session=len(history)
    load_act=load_activity(current_user)
    timedat=datetime.now().strftime('%d/%m/%Y, %H:%M')
    if session>=5:
        badges.append("🥉Beginner Level💫")
        
        load_act.append({
        "time":timedat,
        "activity":f"Achieved 🥉Beginner Level💫 badge"
    })
        save_activity(load_act,current_user)
    if session>=25:
        badges.append("🥈Consistent student✅")
        load_act.append({
        "time":timedat,
        "activity":f"Achieved 🥈Consistent student✅ badge"
    })
        save_activity(load_act,current_user)
    if session>=100:
        badges.append("🥇Master study ✅💪")
        load_act.append({
        "time":timedat,
        "activity":f"Achieved 🥇Master study ✅💪 badge"
    })
        save_activity(load_act,current_user)

    return badges
def view_badges(current_username):
    badge=get_badges(current_username)
    print(Fore.BLUE+"======== Achievements🏅 ========")
    if not badge:
        print(Fore.RED+"Badges not earned yet")

    for badges in badge:
        print(badges)

def weekly_report(current_username):
    history=load_study_history(current_username)
    sessions=0
    min=0
    today=datetime.now().date()
    daybefweek=today-timedelta(days=7)
    strdaybef=daybefweek.strftime("%d/%m/%Y")
    for his in history:
       if his['date']>strdaybef:
        min+=his['duration']
        sessions+=1
    print(Fore.BLUE+"====== Weekly report ======")
    print(f"{Fore.YELLOW}Total sessions :{sessions}")
    print(f"Study time: {min}mins")
def achievements():
    global current_user
    while(True):
        print(Fore.MAGENTA+"1. View badges")
        print(Fore.MAGENTA+"2. Weekly report")
        print(Fore.MAGENTA+"3. Back")
        choice=input(Fore.CYAN+"Enter your choice")
        if choice=='1':
            view_badges(current_user)
        elif choice=='2':
            weekly_report(current_user)
        elif choice=="3":
            break
        else:
            print(f"{Fore.RED}Invalid choice")
def recent_activities(current_username):
    load=load_activity(current_username)
    today=datetime.now().date().strftime('%d/%m/%Y')
    
    start = max(0, len(load) - 5)

    while start < len(load):
      print(Fore.YELLOW+load[start]['activity'])
      start += 1

    
def search_activity(current_username):
    load=load_activity(current_username)
    search=input("Enter which activity title or date(DD/MM/YYYY) you want to search : ")
    for act in load:
        if search in act['activity'] or search in act['time']:
            print(f"Time: {act['activity']}\n Activity: {act['activity']}")
def activity_stats(current_username):
    load=load_activity(current_username)
    study_sessions=0
    viewed_study_history=0
    added_asssignment=0
    viewed_assignment=0
    completed_assignments=0
    added_note=0
    deleted_note=0
    exported_notes=0
    assignment_exported=0
    exported_study_history=0
    tasks_exported=0
    goals_added=0
    viewed_goals=0
    pomodoro=0
    viewed_progress=0
    login=0
    changed_pass=0
    changed_username=0
    added_task=0
    viewed_task=0
    task_deleted=0
    pdf_report=0
    impo_task=0
    impo_notes=0
    stud_impo=0
    study_graph=0
    assignment_chart=0
    excel_report=0
    backup_created=0
    backup_restored=0
    for act in load:
        if act['activity']=='Started study session':
            study_sessions+=1
        elif act['activity']=='Viewed study history':
            viewed_study_history+=1
        elif 'Added assignment' in act['activity']:
            added_asssignment+=1
        elif act['actiity']=='Viewed assignments':
            viewed_assignment+=1
        elif "Completed assignment" in act['activity']:
            completed_assignments+=1
        elif act['activity']=='Note Added':
            added_note+=1
        elif act['activity']=='Note Deleted':
            deleted_note+=1
        elif act['activity']=="Exported notes":
            exported_notes+=1
        elif act['activity']=="Assignment Exported":
            assignment_exported+=1
        elif act['activity']=="Exported Study history":
            exported_study_history+=1
        elif act['activity']=="Tasks Exported":
            tasks_exported+=1
        elif act['activity']=="Added Goal":
            goals_added+=1
        elif act['activity']=="Viewed Goals":
            viewed_goals+=1
        elif act['activity']=="Started Pomodoro":
            pomodoro+=1
        elif act['activity']=="Viewed Progress":
            viewed_progress+=1
        elif act['activity']=="logined in account":
            login+=1
        elif act['activity']=="Changed password":
            changed_pass+=1
        elif "Task added" in act['activity']:
            added_task+=1
        elif act['activity']=="Viewed tasks":
            viewed_task+=1
        elif "task deleted" in act['activity']:
            task_deleted+=1
        elif act['activity']=="Pdf report generated":
            pdf_report+=1
        elif act['activity']=="Tasks Imported":
            impo_task+=1
        elif act['activity']=="Notes imported":
            impo_notes+=1
        elif act['activity']=="Assignments imported":
            assign_impo+=1
        elif act['activity']=="Study history imported":
            stud_impo+=1
        elif act['activity']=="Study graph uploaded":
            study_graph+=1
        elif act['activity']=="Uploaded assignment chart":
            assignment_chart+=1
        elif act['activity']=="Excel report generated":
            excel_report+=1
        elif act['activity']=="Backup created":
            backup_created+=1
        elif act['activity']=="Backup restored":
            backup_restored+=1
        elif act['activity']=="Changed username":
            changed_username+=1
    print(f"study sessions :  {study_sessions}")
    print(f"Viewed study history:   {viewed_study_history}")
    print(f"Added assignments  :  {added_asssignment}")
    print(f"Viewed assignments :  {viewed_assignment}")
    print(f"completed assignments :  {completed_assignments}")
    print(f"study sessions :  {study_sessions}")
    print(f"Added notes :  {added_note}")
    print(f"Deleted notes :  {deleted_note}")
    print(f"Exported notes :  {exported_notes}")
    print(f"Assugnment exported :  {assignment_exported}")
    print(f"Exported study history :  {exported_study_history}")
    print(f"Tasks exported :  {tasks_exported}")
    print(f"Goals added :  {goals_added}")
    print(f"Viewed goals :  {viewed_goals}")
    print(f"Pomodoro sessions :  {pomodoro}")
    print(f"viewed progress:  {viewed_progress}")
    print(f"logined :  {login}")
    print(f"Changed password :  {changed_pass}")
    print(f"changed username: {changed_username}")
    print(f"added task :  {added_task}")
    print(f"task deleted :  {task_deleted}")
    print(f"Pdf reports :  {pdf_report}")
    print(f"import tasks :  {impo_task}")
    print(f"import notes :  {impo_notes}")
    print(f"study imports:  {stud_impo}")
    print(f"Study graphs generated: {study_graph}")
    print(f"Assignment chart generated: {assignment_chart}")
    print(f"Excel report generated : {excel_report}")
    print(f"Backup created: {backup_created}")
    print(f"Backup restored: {backup_restored}")

def activity_menu():
    global current_user
    while(True):
        print(f"{Fore.MAGENTA}1. Search activity")
        print(f"{Fore.MAGENTA}2. Check activity statistics")
        print(f"{Fore.MAGENTA}3. Back")
        choice=input(Fore.CYAN+"Enter your choice : ")
        if choice=='1':
            search_activity(current_user)
        elif choice=='2':
            activity_stats(current_user)
        elif choice=='3':
            break
        else:
            print(Fore.RED+"Invalid choice")

        
        
        
        
        









    




# recent_activities('krishna')
def dashboard():
    global current_user
    load=load_Settings(current_user)
    dat=datetime.now().date()
    today=datetime.now().date()
    print(f"""\n{Fore.BLUE}╔════════════════════════════════════════════════════════════════════╗
║                   🎓 AI STUDENT ASSISTANT                         ║
║                     Productivity Dashboard                        ║
╚════════════════════════════════════════════════════════════════════╝""")
    ti=datetime.now().time().strftime("%H:%M")
    if ti>'12:00':
        greet="👋 Good Afternoon"
    elif ti>'17:00':
        greet="🌙 Good Evening"
    elif ti<'12:00':
        greet="☕ Good Morning"
    print(f"{Fore.YELLOW}{greet}, {Fore.WHITE}{load[0]['display_name']}!")
    print(f"{Fore.YELLOW}📅 Date : {Fore.WHITE}{dat.strftime('%A, %B %d, %Y')}    {Fore.YELLOW}🕒 Time: {Fore.WHITE}{ti}\n{Fore.YELLOW}🔥 Streak: {Fore.WHITE} {view_streak(current_user)} days    {Fore.YELLOW}Productivity score: {Fore.WHITE}{productivity_score(current_user)}")
    print(f"═══════════════════════════════════════════════════════════════════════\n")
    print(f"📊 STUDY PROGRESS")
    print(f"{Fore.YELLOW}Today's goal: {Fore.WHITE}{view_goal(current_user)} Minutes ")
    load_stud=load_study_history(current_user)
    today_study_time=0
    for study in load_stud:
        if study['date']==today.strftime("%d/%m/%Y"):
            today_study_time+=study['duration']
    print(f"{Fore.YELLOW}Studied today: {Fore.WHITE}{today_study_time}")
    percent=(today_study_time / int(view_goal(current_user))) * 100 if view_goal(current_user) != Fore.RED+"please add goal..." else 0
    print(f"{Fore.YELLOW}Progress: {Fore.WHITE}{progress_bar(percent)} {percent:.1f}%\n")
    print(f"{Fore.YELLOW}Total Study Sessions: {Fore.WHITE}{total_study_sessions(current_user)}")
    print(f"{Fore.YELLOW}Total Study Time: {Fore.WHITE}{total_study_time(current_user)} minutes\n")
    
    

  
    # if pending_assignments()<1:
    #     print(f"{Fore.GREEN}Great job! You have no pending assignments.")
    # elif pending_assignments()==total_assignments_count():
    #     print(f"{Fore.RED}You have a lot of pending assignments. Try to complete them soon!")
    print(f"═══════════════════════════════════════════════════════════════════════\n")
    print(f"📋 ASSIGNMENT TRACKING")
    print(f"{Fore.YELLOW}Total Assignments: {Fore.WHITE}{total_assignments_count(current_user)}")
    print(f"{Fore.YELLOW}Pending Assignments: {Fore.WHITE}{pending_assignments(current_user)}")
    print(f"{Fore.GREEN}Completed Assignments: {Fore.WHITE}{completed_assignments(current_user)}")
    print(f"{Fore.RED}Overdue Assignments: {Fore.WHITE}{overdue_count(current_user)}")

    completion_rate= (completed_assignments(current_user) / total_assignments_count(current_user) * 100) if total_assignments_count(current_user) > 0 else 0
    print(f"{Fore.CYAN}Assignment Completion Rate: {progress_bar(completion_rate)} {percent:.1f}%%")

    if completion_rate>=80:
        print(Fore.GREEN+"Excellent work! Your assignment completion rate is high.")
    print(f"═══════════════════════════════════════════════════════════════════════\n")
    print(f"📝 TASK MANAGEMENT")
    total_taskss=total_tasks(current_user)
    print(f"{Fore.YELLOW}Total Tasks: {Fore.WHITE}{total_tasks(current_user)} ")
    print(f"{Fore.YELLOW}Completed Tasks: {Fore.WHITE}{completed_tasks(current_user)} ")
    print(f"{Fore.YELLOW}Pending Tasks: {Fore.WHITE}{pending_tasks(current_user)} ")
    print(f"{Fore.YELLOW}High priority tasks: {Fore.WHITE}{high_priority_tasks(current_user)} 🔴 ")
    print(f"{Fore.YELLOW}Medium priority tasks: {Fore.WHITE}{medium_priority_tasks(current_user)} 🟡 ")
    print(f"{Fore.YELLOW}low priority tasks: {Fore.WHITE}{low_priority_tasks(current_user)} 🟢 ")
    if total_taskss==0:
        print(f"{Fore.RED}Tasks not added yet ")

        
    completed_tas=completed_tasks(current_user)
    total_task=total_tasks(current_user)
    if total_task==0:
        print(Fore.RED+"...")
    else:
      task_completion_rate=(completed_tas/total_task)*100
      if task_completion_rate<=40:
        print(Fore.RED+"You have lots of pending assignments")
      elif task_completion_rate>40 and task_completion_rate<=69:
        print(Fore.YELLOW+"You're making progress")
      elif task_completion_rate>69 and task_completion_rate<=85:
        print(Fore.BLUE+"Great consistency 💫")
      elif task_completion_rate>=85 and task_completion_rate<=100:
        print(Fore.GREEN+"Excellent!, you're completing almost every task✅")
    print(f"═══════════════════════════════════════════════════════════════════════\n")
    print(f"📈 PRODUCTIVITY ")
    print(f"{Fore.YELLOW}Badges Earned: {Fore.WHITE}{', '.join(get_badges(current_user)) if get_badges(current_user) else 'No badges earned yet'}")
    print(f"{Fore.YELLOW}Pomodoro sessions completed: {Fore.WHITE}{load_pomodoro(current_user)[0]['pomodoro']}")
    print(f"{Fore.YELLOW}Productivity score: {Fore.WHITE}{productivity_score(current_user)}")
    print(f"═══════════════════════════════════════════════════════════════════════\n")
    print(f"Recent Activity")
    recent_activities(current_user)

    print(f"═══════════════════════════════════════════════════════════════════════\n")
    print(f"💡  RECOMMENDATIONS")
    recommendation(current_user)

    print(f"═══════════════════════════════════════════════════════════════════════\n")
    print(f"🌟 MOTIVATIONAL QUOTE")
    print(f"{Fore.WHITE}\"{Fore.YELLOW}{motivational_quotes()}{Fore.WHITE}\"")
    print("- Anonymous")





def start_assistant():
   global current_user

   print(f"{Fore.YELLOW}=========== Student Assistant ==========")
   print(f"{Fore.GREEN} ====== Welcome {current_user} ====== ")
   
   while(True):    
        print(f"\n{Fore.BLUE}Choose any option ")
        print(f"{Fore.MAGENTA}1. Study timer")
        print(f"{Fore.MAGENTA}2. Notes")
        print(f"{Fore.MAGENTA}3. View study history")
        print(f"{Fore.MAGENTA}4. Assignments")
        print(f"{Fore.MAGENTA}5. Tasks")
        print(f"{Fore.MAGENTA}6. Dashboard")
        print(f"{Fore.MAGENTA}7. Alerts")
        print(f"{Fore.MAGENTA}8. Export Data")
        print(f"{Fore.MAGENTA}9. Import Data")

        print(f"{Fore.MAGENTA}10. Settings")
        print(f"{Fore.MAGENTA}11. Search")
        print(f"{Fore.MAGENTA}12. Productivity features")
        print(f"{Fore.MAGENTA}13. Insights")
        print(f"{Fore.MAGENTA}14. Reports")
        print(f"{Fore.MAGENTA}15. View achievements")
        print(f"{Fore.MAGENTA}16. View Profile")
        print(f"{Fore.MAGENTA}17. Backup System")
        print(f"{Fore.MAGENTA}18. Logout")
        print(f"{Fore.RED}19. Exit")
        choice=input(Fore.CYAN+"Enter choice : ")
        if choice=='1':
            study_timer(current_user)
        elif choice=='2':
            notes_menu()
        elif choice=='3':
            view_study_history(current_user)
        elif choice=='4':
            assignment_menu()
        elif choice=='5':
            tasks(current_user)
        elif choice=='6':
            dashboard()
        elif choice=='7':
            alerts_menu()
        elif choice=='8':
            export_menu()
        elif choice=='10':
            result=setting_menu()
            if result=="logout":
                return                
        elif choice=='11':
            search_menu()
        elif choice=='12':
            productivity_menu()
        elif choice=='13' :
            insights_menu()
        elif choice=='14':
            reports_menu()
        elif choice=='15':
            achievements()
        elif choice=='16':
            profile()
        elif choice=='17':
            backup_menu()
            break
        elif choice=='18':
           logout()
        elif choice=='19':
            print(Fore.BLUE+"Exiting student assistant...")
            return False
        else:
            print(Fore.RED+"INVALID CHOICE ")
